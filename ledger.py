from collections import namedtuple
from copy import copy
from datetime import datetime
import re
import sqlite3
import openpyxl
import math
import json
from openpyxl import Workbook
import pdfkit

def ceil(no): 
    return math.ceil(no)
def my_round(n, ndigits=0):
    part = n * 10 ** ndigits
    delta = part - int(part)
    # always round "away from 0"
    if delta >= 0.5 or -0.5 < delta <= 0:
        part = math.ceil(part)
    else:
        part = math.floor(part)
    return part / (10 ** ndigits)

class Ledger(object):
    ACCOUNT_TYPES = ('asset', 'liability', 'equity', 'revenue', 'expense')


    def __init__(self, database):
        self.db = database

    def init(self):
        '''Initialize the database.'''
        self.db.executescript('''
        CREATE TABLE IF NOT EXISTS accounts(
            code VARCHAR(255) PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            type VARCHAR(255) NOT NULL,
            balance INTEGER NOT NULL,
			phone VARCHAR(255) ,
			email VARCHAR(255) 
        );
		CREATE TABLE IF NOT EXISTS accounts_dtls(
            code VARCHAR(255) PRIMARY KEY,
            balance INTEGER NOT NULL,
			area INTEGER,
			occupancy INTEGER,
			p2w INTEGER,
			p4w INTEGER,
			dueamount INTEGER NOT NULL,
			dueinterest INTEGER NOT NULL
        );
        CREATE TABLE IF NOT EXISTS transactions(
            id INTEGER PRIMARY KEY,
            date VARCHAR(255) NOT NULL,
            description VARCHAR(255) NOT NULL
        );
        CREATE TABLE IF NOT EXISTS transaction_items(
            id INTEGER PRIMARY KEY,
            transaction_id INTEGER NOT NULL REFERENCES transactions(id),
            account_code VARCHAR(255) NOT NULL REFERENCES accounts(code),
            amount INTEGER NOT NULL,
			type VARCHAR(255) NOT NULL
        );
        ''')
        self.db.commit()

    def generate_bills_excel(self,today):
        #print('Today ->'+today)
        book = Workbook()
        sheet = book.active
        sheet.append(range(1, 21))
        #today = datetime.today()
        date = datetime(today.year, today.month, 1)
        print(ceil(10.1))
        billno = 2001
        i = 1
        bill_conf = self.read_bills()
        bill_items = []
        headers = []
        cursor = self.db.execute(''' SELECT a.code, name, area, a.balance,occupancy,p2w,p4w,dueamount,dueinterest  from accounts a, accounts_dtls ad 
					where a.code = ad.code and 
					ad.area != 0''')

        for code,name,area,balance,occupancy,p2w,p4w,dueamount,dueinterest in cursor:
            a= []
            a.append(date)
            a.append(billno)
            headers.append('Month')
            headers.append('Bill No.')
            billno = billno + 1
            for conf in bill_conf :
               print(conf.values())
               headers.append(conf['name'])
               a.append(eval(conf['expression']))
            if i==1:
               sheet.append(headers)
               i=i+1
            sheet.append(a)
            print ("CODE = ", code , "NAME = ",name ,"AREA = ", area,"BALANCE = ", balance  )
        filename = 'documents\BILLSUMMERY'+today.strftime('%m%Y')+'.xlsx'
        book.save(filename)
        print ("Operation done successfully")
        return filename

    def generate_bills_pdf(self,today):
        # Give the location of the file
        my_path = 'documents\BILLSUMMERY'+today.strftime('%m%Y')+'.xlsx'
        print(my_path)
        my_wb_obj = openpyxl.load_workbook(my_path)
        my_sheet_obj = my_wb_obj.active
        my_row = my_sheet_obj.max_row
        css = ''
        html = ''
        filename = ''
        billdate = today.strftime('%d-%m-%Y')
        options = {
            'page-size': 'A4',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
        }
        with open('style.css', 'r') as f:
            css = f.read()
        with open('test.html', 'r') as f:
            html = f.read()
          

        for i in range(3, my_row + 1):
            filename = 'documents/'+today.strftime('%m%Y')+'/'+my_sheet_obj.cell(row = i, column = 3).value+'.pdf'
            print(filename)
            pdfkit.from_string(css + html.format( 
              billno=my_sheet_obj.cell(row = i, column = 2).value, 
              billdate=billdate, 
              from_='01-04-2020 ', 
              to='30-04-2020 ', 
              duedate='15-04-2020 ', 
              name=my_sheet_obj.cell(row = i, column = 5).value, 
			  flatno=my_sheet_obj.cell(row = i, column = 3).value, 
              maintenance=my_sheet_obj.cell(row = i, column = 6).value, 
              sinkingfund=my_sheet_obj.cell(row = i, column = 7).value, 
              nonoccupancy=my_sheet_obj.cell(row = i, column = 9).value,  
              servicecharges=my_sheet_obj.cell(row = i, column = 8).value, 
              p2w=my_sheet_obj.cell(row = i, column = 10).value, 
              p4w=my_sheet_obj.cell(row = i, column = 11).value, 
              total=my_sheet_obj.cell(row = i, column = 13).value, 
              arrears=my_sheet_obj.cell(row = i, column = 15).value, 
              interest=my_sheet_obj.cell(row = i, column = 14).value, 
              grandtotal=my_sheet_obj.cell(row = i, column = 16).value),filename, options=options)
        return filename

    def upload_account(self):
        # Give the location of the file
        my_path = "D:\documents\REVISEDBILLSUMMERYRAMAIAPRIL20.xlsx"
        my_wb_obj = openpyxl.load_workbook(my_path)
        my_sheet_obj = my_wb_obj.get_sheet_by_name('accounts')
        my_row = my_sheet_obj.max_row
        accounts_arr = {}
        for i in range(3, my_row + 1):
            code = my_sheet_obj.cell(row = i, column = 1).value
            if code is None:
                break
            name = my_sheet_obj.cell(row = i, column = 2).value
            type = my_sheet_obj.cell(row = i, column = 3).value
            phone = my_sheet_obj.cell(row = i, column = 4).value
            email = my_sheet_obj.cell(row = i, column = 5).value
            balance = my_sheet_obj.cell(row = i, column = 6).value
            area = my_sheet_obj.cell(row = i, column = 7).value
            occupancy = my_sheet_obj.cell(row = i, column = 8).value
            p2w = my_sheet_obj.cell(row = i, column = 9).value
            p4w = my_sheet_obj.cell(row = i, column = 10).value
            dueamount = my_sheet_obj.cell(row = i, column = 11).value
            dueinterest=my_sheet_obj.cell(row = i, column = 12).value


            if occupancy > 0: occupancy = 1
            if p2w is None: p2w = 0
            if p4w is None: p4w = 0
            if p2w>0: p2w = p2w/50
            if p4w>0: p4w = p4w/100


            self.create_account(code, name, type, balance,phone,email,area, occupancy, p2w, p4w, dueamount ,dueinterest)

    def save_bills(self , rjson):
        with open("bills.json", "w") as outfile: 
            outfile.write(rjson)

    def read_bills(self ):
        with open('bills.json', 'r') as openfile: 
            json_object = json.load(openfile) 

        return json_object
	

    def monthly_bills(self,filename):
        # Give the location of the file
        my_path = filename
        my_wb_obj = openpyxl.load_workbook(my_path)
        my_sheet_obj = my_wb_obj.active
        my_row = my_sheet_obj.max_row
        txs = {}	
        for i in range(3, my_row + 1):
            date = my_sheet_obj.cell(row = i, column = 1).value
            code = my_sheet_obj.cell(row = i, column = 3).value
            if code is None:
                break
            #date = datetime.strptime('2020-04-01', '%Y-%m-%d').date()
            #date = my_sheet_obj.cell(row = i, column = 3).value
            description = 'maintenance bill'
            interest = my_sheet_obj.cell(row = i, column = 14).value
            if interest is None:
                interest = 0
            amount = my_sheet_obj.cell(row = i, column = 13).value  + interest
            items = []
            items.append(('CORE-ACCOUNT', -amount))
            items.append((code, amount))
            print(code , amount)
            self.record_transaction(date, description, items)

    def bills_payment(self,filename):
        # Give the location of the file
        my_path = filename
        my_wb_obj = openpyxl.load_workbook(my_path)
        my_sheet_obj = my_wb_obj.get_sheet_by_name('transactions')
        my_row = my_sheet_obj.max_row
        txs = {}
        for i in range(3, my_row + 1):

            credit = my_sheet_obj.cell(row = i, column = 4).value
            amount = my_sheet_obj.cell(row = i, column = 2).value
            date = my_sheet_obj.cell(row = i, column = 1).value


            description = my_sheet_obj.cell(row = i, column = 3).value
            debit = my_sheet_obj.cell(row = i, column = 5).value
            items = []
            items.append((credit, -amount))
            items.append((debit, amount))
            print(credit , amount)
            self.record_transaction(date, description, items)

    def add_transaction(self, sheetname):
        # Give the location of the file
        my_path = "D:\documents\REVISEDBILLSUMMERYRAMAIAPRIL20.xlsx"
        my_wb_obj = openpyxl.load_workbook(my_path)
        my_sheet_obj = my_wb_obj.get_sheet_by_name(sheetname)
        my_row = my_sheet_obj.max_row
        txs = {}
        for i in range(3, my_row + 1):

            date = my_sheet_obj.cell(row = i, column = 1).value
            craccount = my_sheet_obj.cell(row = i, column = 2).value
            draccount = my_sheet_obj.cell(row = i, column = 3).value

            amount = my_sheet_obj.cell(row = i, column = 5).value
            description = my_sheet_obj.cell(row = i, column = 4).value

            items = []
            items.append((craccount, amount))
            items.append((draccount, -amount))
            print(craccount , draccount, amount)
            self.record_transaction(date, description, items)

    def drop(self):
        '''Reset the ledger.'''
        self.db.executescript('''
        DROP TABLE IF EXISTS transaction_items;
        DROP TABLE IF EXISTS transactions;
        DROP TABLE IF EXISTS accounts;
		DROP TABLE IF EXISTS accounts_dtls;
        ''')
        self.db.commit()

    def reset(self):
        '''Reset the ledger.'''
        self.drop()
        self.init()

    def create_account(self, code, name, type, balance=0, phone='', email='',area=0, occupancy=0, p2w=0, p4w=0, dueamount=0,dueinterest=0):
        '''Create an account with a given code and name.'''
        if type not in self.ACCOUNT_TYPES:
            raise ValueError('unknown account type {}'.format(type))
        if self.get_account(code):
            raise LedgerError('The account "{}" already exists'.format(code))
        self.db.execute(
            'INSERT INTO accounts(code, name, type, balance,phone,email) VALUES (?, ?, ?, ?,?,?)',
            (code, name, type , balance, phone, email)
        ).close()
        print(code, balance, area,occupancy,p2w,p4w,dueamount,dueinterest)
        self.db.execute(
            'INSERT INTO accounts_dtls(code, balance, area,occupancy,p2w,p4w,dueamount,dueinterest) VALUES (?, ?, ?,?, ?, ?,?,?)',
            (code, balance, area,occupancy,p2w,p4w,dueamount,dueinterest)
        ).close()
		
        self.db.commit()

    def edit_account(self, code, name, type, balance=0, phone='', email=''):
        '''Create an account with a given code and name.'''
        if type not in self.ACCOUNT_TYPES:
            raise ValueError('unknown account type {}'.format(type))
        if self.get_account(code) is not None:
            raise LedgerError('The account "{}" not  exists'.format(code))
        self.db.execute(
            'UPDATE accounts name=?,phone=?, email=? WHERE code =?',
            (name, phone , email, code)
        ).close()
        self.db.commit()

    def get_balance_sheet(self, date):
        '''Return a balance sheet.'''
        print(date)
        rows = self.db.execute('''
         SELECT
            a.code, a.name, a.type, SUM(
                CASE
                    WHEN ti.amount is NOT NULL  THEN ti.amount+ad.balance
                    ELSE ad.balance
                END
            ) as bal
            FROM accounts a
            LEFT JOIN transaction_items ti ON a.code = ti.account_code
            LEFT JOIN transactions t ON ti.transaction_id = t.id 
            AND date(t.date) <= date(?)
            LEFT JOIN accounts_dtls ad on a.code=ad.code
            GROUP BY a.code
        ''', (date,)).fetchall()

        retained_earnings = 0
        accounts_by_type = {'asset': {}, 'liability': {}, 'equity': {}}
        for code, name, type, balance in rows:
            if type in ('revenue', 'expense'):
                retained_earnings -= balance
            else:
                accounts_by_type[type][Account(code, name,'','','', type)] = balance

        return BalanceSheet(
            date=date,
            retained_earnings=retained_earnings,
            **accounts_by_type
        )

    def get_income_statement(self, start_date, end_date):
        '''Return an income statement.'''

        rows = self.db.execute('''
        SELECT
            a.code, a.name, a.type, SUM(
                CASE
                    WHEN a.type ="revenue" 
                    	THEN CASE
                    			WHEN ti.type ="2" AND ti.amount>0
                        			THEN ti.amount
                        			ELSE 0
                        		END
                    ELSE ti.amount
                END
            ) as balance
            FROM accounts a
            LEFT JOIN transaction_items ti ON a.code = ti.account_code
            INNER JOIN transactions t ON ti.transaction_id = t.id 
            AND date(?) <= date(t.date) AND date(t.date) <= date(?)
            WHERE a.type IN ("revenue", "expense")
            GROUP BY a.code
        ''', (start_date, end_date,)).fetchall()

        accounts_by_type = {'revenue': {}, 'expense': {}}
        for code, name, type, balance in rows:
            if balance is not None:
                accounts_by_type[type][Account(code, name,'','','', type)] = balance

        return IncomeStatement(
            start_date=start_date,
            end_date=end_date,
            **accounts_by_type
        )

    def get_account(self, code):
        '''Return the account identified by the specified code.'''
        row = self.db.execute('SELECT * FROM accounts WHERE code = ?',
                              (code,)).fetchone()
        if row is None:
            return None
        return Account(row[0], row[1], row[2], row[3],row[4],row[5])


    def get_allaccount(self):
        '''Return the account identified by the specified code.'''
        accounts_arr = {}
        rows = self.db.execute('SELECT * FROM accounts').fetchall()
        i = 0
        if rows is None:
            return None
        for code,name,type,balance,phone,email in rows:
            accounts_arr[i] = Account(code, name, phone,email,balance,type)
            i = i + 1
        return accounts_arr.values()    


    def record_transaction(self, date, description, items):
        '''Record a transaction.'''
        if not items:
            raise ValueError('cannot record an empty transaction')
        print (items)
        if sum(item[1] for item in items) != 0:
            raise ValueError('unbalanced transaction items')

        try:
            c = self.db.cursor()
            c.execute(
                'INSERT INTO transactions(date, description) VALUES (?, ?)',
                (date.strftime('%Y-%m-%d'), description)
            )
            tx_id = c.lastrowid
            index = 1
            for account_code, amount in items:
                c.execute(
                    'SELECT 1 FROM accounts WHERE code = ?',
                    (account_code,)
                )
                if c.fetchone() is None:
                    raise ValueError(
                        'unknown account code {}'.format(account_code)
                    )

                c.execute('''INSERT INTO transaction_items(transaction_id,
                                                           account_code,
                                                           amount,type)
                                                           VALUES (?, ?, ?, ?)''',
                          (tx_id, account_code, amount, index))
                c.execute(
                    'UPDATE accounts SET balance=(balance + ?) WHERE code = ?',
                    (amount,account_code)
                )
                index += 1
        except:
            self.db.rollback()
            raise
        finally:
            c.close()

        self.db.commit()
        return tx_id

    def count_transactions(self):
        '''Return the number of transactions.'''
        return self.db.execute(
            'SELECT COUNT(*) FROM transactions'
        ).fetchone()[0]

    def count_transaction_items(self):
        '''Return the number of transaction items.'''
        return self.db.execute(
            'SELECT COUNT(*) FROM transaction_items'
        ).fetchone()[0]

    def get_transactions(self):
        '''Return all registered transactions.'''
        txs = {}

        rows = self.db.execute('SELECT * FROM transactions order by date desc').fetchall()
        for tx_id, date, description in rows:
            date = datetime.strptime(date, '%Y-%m-%d').date()
            txs[tx_id] = Transaction(id, date, description, [])

        rows = self.db.execute('SELECT * FROM transaction_items').fetchall()
        for _, tx_id, account_code, amount, type in rows:
            txs[tx_id].items.append((account_code, amount, type))

        return txs.values()

    def get_accountledger(self, account_code):
        '''Return all registered transactions.'''
        ldg = {}
        cnt = 1
        balance = 0
        rows = self.db.execute('''SELECT ti.transaction_id,  ti.amount, 
								ti.type, t.date, t.description,	a.balance FROM transaction_items  ti
								INNER JOIN transactions t ON ti.transaction_id = t.id
								INNER JOIN accounts_dtls a ON ti.account_code = a.code
								where ti.account_code = ?  order by t.date''',(account_code,)).fetchall()
        for l_id, amount, type, date, description, g_balance in rows:
            date = datetime.strptime(date, '%Y-%m-%d').date()
            if cnt == 1:
                balance = g_balance
                cnt = cnt + 1
            balance = balance + amount
            ldg[l_id] = LedgerItems(l_id, date, description, amount, type, balance)

        return ldg.values()

    def get_transaction(self, tx_id):
        '''Return the specified transaction.'''
        row = self.db.execute('SELECT * FROM transactions WHERE id = ?',
                              (tx_id,)).fetchone()
        if row is None:
            return None

        date = datetime.strptime(row[1], '%Y-%m-%d').date()
        description = row[2]

        items = []
        rows = self.db.execute(
            'SELECT * FROM transaction_items WHERE transaction_id = ?',
            (tx_id,)
        ).fetchall()
        for row in rows:
            items.append((row[2], row[3], row[4]))

        return Transaction(tx_id,date, description, items)




class LedgerError(RuntimeError):
    pass


Account = namedtuple('Account', 'code name phone email balance type')
Transaction = namedtuple('Transaction', 'id date description items')
LedgerItems = namedtuple('LedgerItems', 'id date description amount type balance')


class BalanceSheet(namedtuple('BalanceSheet',
                              'date asset liability equity retained_earnings')):
    @property
    def total_assets(self):
        return sum(self.asset.values())

    @property
    def total_liabilities(self):
        return -sum(self.liability.values())

    @property
    def total_equity(self):
        return -sum(self.equity.values()) + self.retained_earnings


class IncomeStatement(namedtuple('IncomeStatement',
                                 'start_date end_date revenue expense')):
    @property
    def total_revenues(self):

        return sum(self.revenue.values())

    @property
    def total_expenses(self):

        return sum(self.expense.values())

    @property
    def net_result(self):
        return self.total_revenues - self.total_expenses

    @property
    def net_income(self):
        if self.net_result >= 0:
            return self.net_result
        return 0

    @property
    def net_loss(self):
        if self.net_result < 0:
            return - self.net_result
        return 0
